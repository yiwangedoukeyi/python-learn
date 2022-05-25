import datetime
import yaml
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment

Config = {}
StartDate = datetime.datetime.today()
EndDate = datetime.datetime.today()
HeaderSize = 2


def load_cfg():
    global Config
    with open('conf.yaml', 'r', encoding='utf-8') as f:
        cfg_raw = f.read()
        Config = yaml.safe_load(cfg_raw)


sequence = list(map(lambda x: chr(x), range(ord('A'), ord('Z') + 1)))


def get_excel_col_index(num):
    num -= 1
    col_list = []
    if num > 25:
        while True:
            d = int(num / 26)
            remainder = num % 26
            if d <= 25:
                col_list.insert(0, sequence[remainder])
                col_list.insert(0, sequence[d - 1])
                break
            else:
                col_list.insert(0, sequence[remainder])
                num = d - 1
    else:
        col_list.append(sequence[num])

    return "".join(col_list)


def excel_merge_and_center(ws, row, column, value=None, merge_str=None):
    cell = ws.cell(row=row, column=column)
    if value is not None:
        cell.value = value
    if merge_str is not None and merge_str != '':
        ws.merge_cells(merge_str)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def write_excel_header(ws):
    excel_merge_and_center(ws, 1, 1, "日期", 'A1:A2')
    ws.column_dimensions['A'].width = 12

    column_index = 2
    plan_start_col_indexes = []
    for cfg in Config["plan"]:
        plan_start_col_indexes.append(column_index)
        cells_num = 0
        day_count = cfg["daily_count"]
        excel_merge_and_center(ws, 1, column_index, cfg["plan_name"])
        for i in range(day_count):
            count_str = '第' + str(i + 1) + '次'
            cells_num += 1
            excel_merge_and_center(ws, 2, column_index + i, count_str)
            ws.column_dimensions[get_excel_col_index(column_index + i)].width = 15
        if cells_num > 1:
            merge_cell_str = (get_excel_col_index(column_index) +
                              '1:' + get_excel_col_index(cells_num - 1 + column_index) + '1')
            excel_merge_and_center(ws, 1, column_index, merge_str=merge_cell_str)
        column_index += cells_num
    return plan_start_col_indexes


def write_one_day_plan_excel(ws, start_row_index, start_col_index, plan_index_list):
    for i in range(len(plan_index_list)):
        ele_num = len(plan_index_list[i]) - 1
        value = ""
        for index in plan_index_list[i]:
            value += str(index) + ','
        value = value.rstrip(',')
        if ele_num > 3:
            ws.column_dimensions[get_excel_col_index(start_col_index)].width = 5 * ele_num
        excel_merge_and_center(ws, start_row_index, start_col_index + i, value)


def gen_day_plan(plan_index_list, cfg, differ):
    min_index = 0
    min_size = len(plan_index_list[min_index])
    for i in range(len(plan_index_list)):
        if len(plan_index_list[i]) < min_size:
            min_size = len(plan_index_list[i])
            min_index = i

        for j in range(len(plan_index_list[i])):
            plan_index_list[i][j] += 1
            if plan_index_list[i][j] > cfg["end_chapter"]:
                del plan_index_list[i][j]
    if differ in Config["rule"]:
        if differ == 1:
            # 第一次放到最下方
            plan_index_list[-1].insert(0, int(cfg["start_chapter"]))
        else:
            plan_index_list[min_index].insert(0, int(cfg["start_chapter"]))


def write_one_plan(ws, cfg, start_col_index):
    plan_start_time = datetime.datetime.strptime(cfg["start_date"], "%Y-%m-%d")
    delta = plan_start_time - StartDate
    start_row_index = delta.days + HeaderSize + 1
    plan_index_list = []
    for i in range(cfg["daily_count"]):
        temp_index_list = [int(cfg["start_chapter"])]
        plan_index_list.append(temp_index_list)

    write_one_day_plan_excel(ws, start_row_index, start_col_index, plan_index_list)

    delta = EndDate - plan_start_time
    day_num = delta.days
    for i in range(1, day_num + 1):
        cur_row_index = start_row_index + i
        gen_day_plan(plan_index_list, cfg, i)
        # print(plan_index_list)
        write_one_day_plan_excel(ws, cur_row_index, start_col_index, plan_index_list)


def write_date(ws):
    global StartDate
    global EndDate
    delta = timedelta(days=int(Config["end_count"]) - 1)
    EndDate = StartDate + delta
    day_num = delta.days

    for i in range(day_num + 1):
        row_index = HeaderSize + i + 1
        delta = timedelta(days=i)
        now_time = StartDate + delta
        excel_merge_and_center(ws, row_index, 1, now_time.strftime('%Y-%m-%d'))


def write_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = Config["file_name"]
    plan_start_col_indexes = write_excel_header(ws)
    write_date(ws)
    for i in range(len(Config["plan"])):
        write_one_plan(ws, Config["plan"][i], plan_start_col_indexes[i])

    wb.save(Config["file_name"] + ".xlsx")


def gen_daily_plan():
    global StartDate
    StartDate = datetime.datetime.strptime(Config["plan"][0]["start_date"], "%Y-%m-%d")
    for cfg in Config["plan"]:
        one_time = datetime.datetime.strptime(cfg["start_date"], "%Y-%m-%d")
        if one_time < StartDate:
            StartDate = one_time
    write_excel()


if __name__ == '__main__':
    load_cfg()
    gen_daily_plan()
